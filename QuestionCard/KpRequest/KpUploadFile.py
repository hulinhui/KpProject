from QuestionCard.KpRequest.KpLogin import KpLogin
from QuestionCard.PdfConvertImage import get_file_path
from openpyxl import Workbook


class KpUploadFile:
    def __init__(self, up_type):
        self.object = KpLogin()
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.up_list = ['stu', 'tea', 'score']
        self.upload_type = self._check_type(up_type)

    def _check_type(self, up_type):
        if not isinstance(up_type, str):
            raise TypeError(f'Type of {up_type} is not str')
        if up_type not in self.up_list:
            raise Exception("上传类型不满足条件")
        return up_type

    def __del__(self):
        self.workbook.close()

    def append_cols(self, cols_list):
        last_column_index = self.sheet.max_column
        for index, col in enumerate(cols_list, 1):
            cols_index = last_column_index + index
            self.sheet.insert_cols(cols_index)
            self.sheet.cell(row=1, column=cols_index).value = col

    def delete_cols(self):
        pass

    def create_excel_data(self):
        pass

    def get_upload_info(self):
        """
        获取上传所需数据
        :return: tuple  上传url、上传data、上传文件路径
        """
        data = self.object.kp_data
        upload_data = {'examId': '040172742128913303800', 'paperId': '040172742153751303814000'}
        folder_path = get_file_path('excel')
        upload_item = {
            'stu': (data['upload_stu_url'], upload_data, get_file_path('参考学生导入模板.xlsx', folder_path)),
            'tea': (data['upload_tea_url'], upload_data, get_file_path('阅卷老师导入模板.xlsx', folder_path)),
            'score': (data['upload_score_url'], upload_data, get_file_path('上传成绩导入模版.xlsx', folder_path))
        }
        return upload_item.get(self.upload_type, "无效的输入")

    def upload_file(self):
        """
        通过文件上传发送请求
        :return: None
        """
        # 请求URL、请求data(同时使用files和data参数，通常会将data也作为一个字典来传递)
        upload_url, upload_data, file_path = self.get_upload_info()
        # 清除Content-Type字段(files参数会自动帮我们加上Content-Type,如果我传了就会覆盖自动添加的，导致请求失败)
        del self.object.headers['Content-Type']
        # 请求文件对象
        with open(file_path, 'rb') as fp:
            upload_response = self.object.get_response(url=upload_url, method='POST', data=upload_data,
                                                       files={'file': fp})
        result, data = self.object.check_response(upload_response)
        if not result:
            # 接口数据返回失败，一种是接口直接返回（errormsg），另一种是记录到错误提示+
            error_data = data['data']
            # error_data存在则输出错误提示，否则输出errormsg信息
            error_data and [self.object.logger.info(f'上传失败-->{i["msg"]}--{i["desc"]}') for i in
                            error_data] or self.object.logger.info(f'上传失败-->{data["errorMsg"]}')
        else:
            # 接口返回成功
            self.object.logger.info(f'上传成功!')

    def run(self):
        # 生成excel数据
        self.create_excel_data()
        # 登录
        self.object.get_login_token()
        # 上传文件
        self.upload_file()


if __name__ == '__main__':
    uf = KpUploadFile(up_type='stu')
    uf.run()
